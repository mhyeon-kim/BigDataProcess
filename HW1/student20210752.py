#!/usr/bin/python3

from openpyxl import load_workbook

wb = load_workbook('./student.xlsx')
ws = wb.active

list_total = []
for row in range(2, ws.max_row + 1):
    midterm = ws['C' + str(row)].value
    final = ws['D' + str(row)].value
    hw = ws['E' + str(row)].value
    attendence = ws['F' + str(row)].value

    total = (midterm * 0.3) + (final * 0.35) + (hw * 0.34) + (attendence * 0.01)

    ws['G' + str(row)].value = total
    list_total.append((total, row))

list_total.sort(reverse = True)
a = int(len(list_total) * 0.3)
ap = a //2
b = int(len(list_total) * 0.7)
bp = (a + b) // 2

for i, (total, row) in enumerate(list_total):
    if i < a:
        ws['H' + str(row)] = 'A'
    elif a <= i < b:
        ws['H' + str(row)] = 'B'
    else:
        ws['H' + str(row)] = 'C'

for i in range(ap):
    row = list_total[i][1]
    ws['H' + str(row)] = 'A+'

for i in range(a, bp):
    row = list_total[i][1]
    ws['H' + str(row)] = 'B+'

for i in range(b, (len(list_total) + b) // 2):
    row = list_total[i][1]
    ws['H' + str(row)] = 'C+'

for row in range(2, ws.max_row + 1):
    if ws['G' + str(row)].value < 40:
        ws['H' + str(row)] = 'F'

wb.save('./student.xlsx')
wb.close()
